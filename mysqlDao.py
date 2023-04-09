import os

import openpyxl
import pymysql
from _testcapi import INT_MAX
from openpyxl.chart import PieChart, Reference, BarChart
from openpyxl.chart.label import DataLabelList

from RecongnizePicture import get_VAT_invoice_context
import config

# 连接数据库，localhost：3306，用户名是root，密码是xgh200011125850，数据库是RPA
conn = pymysql.connect(host='localhost', port=3306, user=config.mysql_username, password=config.mysql_password, db='rpa', charset='utf8')
# 获取游标
cursor = conn.cursor()


# 定义发票信息类，属性为发票信息，包括InvoiceDate，SellerRegisterNum，PurchasserName，SellerName，AmountInFiguers，status
class InvoiceInfo:
    def __init__(self):
        self.InvoiceDate = ""
        self.SellerRegisterNum = ""
        self.PurchasserName = ""
        self.SellerName = ""
        self.AmountInFiguers = ""
        self.status = ""
        self.invoiceName = ""

    # 定义构造函数，传入发票信息
    def __init__(self, InvoiceDate, SellerRegisterNum, PurchasserName, SellerName, AmountInFiguers, status,
                 invoiceName):
        self.InvoiceDate = InvoiceDate
        self.SellerRegisterNum = SellerRegisterNum
        self.PurchasserName = PurchasserName
        self.SellerName = SellerName
        self.AmountInFiguers = AmountInFiguers
        self.status = status
        self.invoiceName = invoiceName

    # 定义get和set函数操作属性
    def getInvoiceDate(self):
        return self.InvoiceDate

    def setInvoiceDate(self, InvoiceDate):
        self.InvoiceDate = InvoiceDate

    def getSellerRegisterNum(self):
        return self.SellerRegisterNum

    def setSellerRegisterNum(self, SellerRegisterNum):
        self.SellerRegisterNum = SellerRegisterNum

    def getPurchasserName(self):
        return self.PurchasserName

    def setPurchasserName(self, PurchasserName):
        self.PurchasserName = PurchasserName

    def getSellerName(self):
        return self.SellerName

    def setSellerName(self, SellerName):
        self.SellerName = SellerName

    def getAmountInFiguers(self):
        return self.AmountInFiguers

    def setAmountInFiguers(self, AmountInFiguers):
        self.AmountInFiguers = AmountInFiguers

    def getStatus(self):
        return self.status

    def setStatus(self, status):
        self.status = status

    def getInvoiceName(self):
        return self.invoiceName

    def setInvoiceName(self, invoiceName):
        self.invoiceName = invoiceName


# 定义一个插入函数，传入发票信息对象，将发票信息插入数据库
def insert_one(invoiceInfo):
    # 定义sql语句
    sql = "insert into vat(InvoiceDate,SellerRegisterNum,PurchasserName,SellerName,AmountInFiguers," \
          "status,invoiceName) values(%s,%s,%s,%s,%s,%s,%s)"
    # 执行sql语句
    cursor.execute(sql, (
        invoiceInfo.InvoiceDate, invoiceInfo.SellerRegisterNum, invoiceInfo.PurchasserName, invoiceInfo.SellerName,
        invoiceInfo.AmountInFiguers, invoiceInfo.status, invoiceInfo.invoiceName))
    # 提交事务
    conn.commit()


# 定义一个批量插入函数，传入datas，datas是一个发票信息对象的列表，将发票信息插入数据库
def insert_many(datas):
    # 定义一个for循环，将datas里信息提取出来构造invoiceInfo对象列表
    amount = INT_MAX
    for d in range(len(datas)):
        if 'InvoiceDate' in datas[d]:
            InvoiceDate = datas[d]['InvoiceDate']
        if 'SellerRegisterNum' in datas[d]:
            SellerRegisterNum = datas[d]['SellerRegisterNum']
        if 'PurchasserName' in datas[d]:
            PurchasserName = datas[d]['PurchasserName']
        if 'SellerName' in datas[d]:
            SellerName = datas[d]['SellerName']
        if 'AmountInFiguers' in datas[d]:
            amount = float(datas[d]['AmountInFiguers'])
            storeFigure = int(amount * 100)
        print('debug:')
        print(datas[d]['InvoiceDate'] == '2016年06月12日')
        print(amount <= 2700.00)
        print(datas[d]['PurchasserName'] == '深圳市购机汇网络有限公司')
        if 'InvoiceDate' not in datas[d] or 'AmountInFiguers' not in datas[d] or 'SellerName' not in datas[d]:
            status = '转人工'
        elif datas[d]['InvoiceDate'] == '2016年06月12日' and amount <= 2700.00 and datas[d][
            'PurchasserName'] == '深圳市购机汇网络有限公司':
            status = '通过'
        else:
            status = '不通过'
        invoiceInfo = InvoiceInfo(InvoiceDate, SellerRegisterNum, PurchasserName, SellerName, storeFigure, status,
                                  datas[d]['invoiceName'])
        # 调用insert_one函数，将发票信息插入数据库
        insert_one(invoiceInfo)
    print("插入mysql成功")


# 定义一个函数，获取一张发票的信息，并插入到数据库
# def get_one_invoice_info(i):
# 定义一个查询函数，查询数据库中所有发票信息并返回
def select_all():
    # 定义sql语句
    sql = "select * from vat"
    # 执行sql语句
    cursor.execute(sql)
    # 获取查询结果
    result = cursor.fetchall()
    # 打印查询结果
    return result


# 定义一个清空表函数，清空数据库中所有发票信息
def delete_all():
    # 定义sql语句
    sql = "delete from vat"
    # 执行sql语句
    cursor.execute(sql)
    # 提交事务
    conn.commit()
    print("清空成功")


# 定义一个函数做数据统计，统计总过有多少条数据
def count_all():
    # 定义sql语句
    sql = "select count(*) from vat"
    # 执行sql语句
    cursor.execute(sql)
    # 获取查询结果
    result = cursor.fetchall()
    # 打印查询结果
    return result[0][0]


# 定义一个函数做数据统计，统计status是通过的有多少条数据
def count_pass():
    # 定义sql语句
    sql = "select count(*) from vat where status='通过'"
    # 执行sql语句
    cursor.execute(sql)
    # 获取查询结果
    result = cursor.fetchall()
    # 打印查询结果
    return result[0][0]


# 定义一个函数做数据统计，统计status是不通过的有多少条数据
def count_not_pass():
    # 定义sql语句
    sql = "select count(*) from vat where status='不通过'"
    # 执行sql语句
    cursor.execute(sql)
    # 获取查询结果
    result = cursor.fetchall()
    # 打印查询结果
    return result[0][0]


# 定义一个函数做数据统计，统计status是转人工的有多少条数据
def count_to_human():
    # 定义sql语句
    sql = "select count(*) from vat where status='转人工'"
    # 执行sql语句
    cursor.execute(sql)
    # 获取查询结果
    result = cursor.fetchall()
    # 打印查询结果
    return result[0][0]


# 定义函数统计共有多少条数据
def count_all():
    # 定义sql语句
    sql = "select count(*) from vat"
    # 执行sql语句
    cursor.execute(sql)
    # 获取查询结果
    result = cursor.fetchall()
    # 打印查询结果
    return result[0][0]


def export_to_excel():
    # 定义sql语句
    sql = "select * from vat order by AmountInFiguers desc"
    # 执行sql语句
    cursor.execute(sql)
    # 获取查询结果
    result = cursor.fetchall()
    # 打印查询结果
    print('result:', result)
    # 提交事务
    conn.commit()
    # 定义一个workbook对象
    wb = openpyxl.Workbook()
    # 定义一个sheet对象
    ws = wb.active
    ws.title = "sheet1"
    # 写入表头
    title = ['发票名称', '开票日期', '纳税人识别号', '购买方名称', '卖方名称', '购买金额', '审批状态', ]
    for i in range(len(title)):
        ws.cell(row=1, column=i + 1, value=title[i])
    # 写入查询结果
    for i in range(len(result)):
        for j in range(len(result[i]) - 2):
            ws.cell(row=i + 2, column=j + 1, value=result[i][j + 1])
    # 写入交易金额和审批状态，交易金额要除以100，并且保留两位小数
    for i in range(len(result)):
        amount = "{:.2f}".format(float(result[i][6]) / 100)
        ws.cell(row=i + 2, column=6, value=amount)
        ws.cell(row=i + 2, column=7, value=result[i][7])

    add_pie_chart(ws, len(result))
    # 保存数据总条数
    ws.cell(row=1, column=8, value='数据总条数')
    ws.cell(row=2, column=8, value=count_all())
    # 保存excel表格
    wb.save(config.excel_filename)


# 添加一个饼图
def add_pie_chart(ws, resultLen):
    # 统计pass、not pass、to human三者的数量
    # 创建饼图对象
    # 统计status是通过的有多少条数据，写入到excel表格中
    ws.cell(row=resultLen + 2, column=1, value='审核状态')
    ws.cell(row=resultLen + 3, column=1, value='passed')
    # 统计status是不通过的有多少条数据，写入到excel表格中
    ws.cell(row=resultLen + 4, column=1, value='not_passed')
    # 统计status是转人工的有多少条数据，写入到excel表格中
    ws.cell(row=resultLen + 5, column=1, value='to_human')
    ws.cell(row=resultLen + 2, column=2, value='审核状态总数')
    ws.cell(row=resultLen + 3, column=2, value=count_pass())
    # 统计status是不通过的有多少条数据，写入到excel表格中
    ws.cell(row=resultLen + 4, column=2, value=count_not_pass())
    # 统计status是转人工的有多少条数据，写入到excel表格中
    ws.cell(row=resultLen + 5, column=2, value=count_to_human())
    pie = PieChart()
    # 设置饼图的数据和标签
    data_ref = Reference(ws, min_col=2, min_row=resultLen + 2, max_col=2, max_row=resultLen + 5)
    print(data_ref)
    pie.add_data(data_ref, titles_from_data=True)
    labels_ref = Reference(ws, min_col=1, min_row=resultLen + 3, max_col=1, max_row=resultLen + 5)
    print(labels_ref)
    pie.set_categories(labels_ref)
    pie.style = 8  # 图表样式类型 共48种
    pie.title = "审核状态统计"  # 图表标题
    pie.height = 10  # 图表高度
    pie.width = 10  # 图表宽度
    s1 = pie.series[0]
    s1.dLbls = DataLabelList()
    s1.dLbls.showCatName = True  # 标签显示
    s1.dLbls.showVal = True  # 数量显示
    s1.dLbls.showPercent = True  # 百分比显示

    ws.add_chart(pie, "G2")


# 定义一个函数，删除导出的excel表格
def delete_excel():
    if os.path.exists(config.excel_filename):
        os.remove(config.excel_filename)


# 定义一个主函数，用于测试
def main():
    # invoiceInfo = InvoiceInfo('2016年06月12日', '91440101664041243T', '深圳市购机汇网络有限公司',
    #                           '广州晶东贸易有限公司', 345000
    #                           , '不通过','b0')
    # insert_one(invoiceInfo)
    # select_all()
    # print('共有', count_all())
    # print('通过的有：', count_pass())
    # print('不通过的有：', count_not_pass())
    # print('转人工的有：', count_to_human())
    pic = config.testPicPath
    datas = []
    data = get_VAT_invoice_context(pic)
    datas.append(data)
    insert_many(datas)
    export_to_excel()
    # delete_all()


# 调用主函数
if __name__ == '__main__':
    main()
