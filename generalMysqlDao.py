# 连接数据库
import os

import openpyxl
import pymysql
from _testcapi import INT_MAX

import config

conn = pymysql.connect(host='localhost', port=3306, user=config.mysql_username, password=config.mysql_password,
                       db='rpa', charset='utf8')
# 获取游标
cursor = conn.cursor()


# 定义general_invoice_info类，属性为发票信息，包括InvoiceDate，PurchasserName，SellerName，AmountInFiguers，status
class general_invoice_info:
    def __init__(self):
        self.InvoiceDate = ""
        self.PurchasserName = ""
        self.SellerName = ""
        self.AmountInFiguers = ""
        self.status = ""
        self.invoiceName = ""

    # 定义构造函数，传入发票信息
    def __init__(self, InvoiceDate, PurchasserName, SellerName, AmountInFiguers, status,
                 invoiceName):
        self.InvoiceDate = InvoiceDate
        self.PurchasserName = PurchasserName
        self.SellerName = SellerName
        self.AmountInFiguers = AmountInFiguers
        self.status = status
        self.invoiceName = invoiceName

    # 定义get和set函数操作属性
    def getInvoiceDate(self):
        return self.InvoiceDate

    def setInvoiceDate(self, InvoiceDate):
        self.InvoiceDate = InvoiceDate

    def getPurchasserName(self):
        return self.PurchasserName

    def setPurchasserName(self, PurchasserName):
        self.PurchasserName = PurchasserName

    def getSellerName(self):
        return self.SellerName

    def setSellerName(self, SellerName):
        self.SellerName = SellerName

    def getAmountInFiguers(self):
        return self.AmountInFiguers

    def setAmountInFiguers(self, AmountInFiguers):
        self.AmountInFiguers = AmountInFiguers

    def getInvoiceName(self):
        return self.invoiceName

    def setInvoiceName(self, invoiceName):
        self.invoiceName = invoiceName

    def getStatus(self):
        return self.status

    def setStatus(self, status):
        self.status = status


# 定义插入函数，传入发票信息
def insertInvoiceInfo(invoiceInfo):
    # 定义sql语句
    sql = "insert into general(InvoiceDate,PurchasserName,SellerName,AmountInFiguers,status,invoiceName) " \
          "values('%s','%s','%s','%s','%s','%s')" % (
              invoiceInfo.InvoiceDate, invoiceInfo.PurchasserName, invoiceInfo.SellerName, invoiceInfo.AmountInFiguers,
              invoiceInfo.status, invoiceInfo.invoiceName)
    # 执行sql语句
    cursor.execute(sql)
    # 提交
    conn.commit()


# 定义批量插入函数，传入发票信息列表
def insertInvoiceInfoList(datas):
    # 定义变量，保存InvoiceDate，PurchasserName，SellerName，AmountInFiguers
    # InvoiceDate = ""
    # PurchasserName = ""
    # SellerName = ""
    # storeFigure = 0
    amount = INT_MAX
    # 定义一个发票信息类
    invoiceInfo = general_invoice_info('1970-01-01', '', '', 0, '', '')
    # 定义一个发票信息类
    # invoiceInfo = general_invoice_info()
    for d in range(len(datas)):
        invoiceInfo.setInvoiceName(datas[d]['invoiceName'])
        status_date = True  # 默认正确识别日期
        # 未识别出结果
        if 'exception' in datas[d]:
            invoiceInfo.setInvoiceDate('1970-01-01')
            invoiceInfo.setPurchasserName("")
            invoiceInfo.setSellerName("")
            invoiceInfo.setAmountInFiguers(0)
            invoiceInfo.setStatus('转人工')
            insertInvoiceInfo(invoiceInfo)
            continue
        # api可以识别，检测识别结果是否有误
        if datas[d]['InvoiceDate']:
            # 如果日期存在的话，将日期转为mysql中的YYYY-MM-DD格式，方便后续查询，而Invoice的格式为YYYY年MM月DD日
            InvoiceDate = datas[d]['InvoiceDate'].replace('年', '-').replace('月', '-').replace('日', '')
            dy = InvoiceDate.split('-')
            if not dy[0]:
                status_date = False
                dy[0] = '1970'
            if not dy[1]:
                status_date = False
                dy[1] = '01'
            if not dy[2]:
                status_date = False
                dy[2] = '01'

            invoiceInfo.setInvoiceDate(dy[0]+'-'+dy[1]+'-'+dy[2])
        else:
            status_date = False
            invoiceInfo.setInvoiceDate('1970-01-01')
        #if 'PurchasserName' in datas[d]:
        if datas[d]['PurchasserName']:
            PurchasserName = datas[d]['PurchasserName']
            invoiceInfo.setPurchasserName(PurchasserName)
        else:
            invoiceInfo.setPurchasserName("")
        # if 'SellerName' in datas[d]:
        if datas[d]['SellerName']:
            SellerName = datas[d]['SellerName']
            invoiceInfo.setSellerName(SellerName)
        else:
            invoiceInfo.setSellerName("")
        # if 'AmountInFiguers' in datas[d]:
        if datas[d]['AmountInFiguers']:
            amount = float(datas[d]['AmountInFiguers'])
            storeFigure = int(amount * 100)
            invoiceInfo.setAmountInFiguers(storeFigure)
        else:
            invoiceInfo.setAmountInFiguers(0)
        # # 如果发票信息中缺少发票日期、金额、付款方的信息，视为转人工
        # if 'InvoiceDate' not in datas[d] or 'AmountInFiguers' not in datas[d] or 'PurchasserName' not in datas[d]:
        #     invoiceInfo.setStatus('转人工')
        # # 如果付款方为"浙江大学"、时间在2015年内、审批金额在1600元以内的发票视为合规
        # elif datas[d]['PurchasserName'] == '浙江大学' and InvoiceDate < '2015-01-01' and amount <= 1600.00:
        #     invoiceInfo.setStatus('通过')
        # else:
        #     invoiceInfo.setStatus('不通过')
        # 调用insert_one函数，将发票信息插入数据库
        # 当年月日有一个没识别出来时，均转人工
        if status_date :
            if  datas[d]['AmountInFiguers'] and datas[d]['PurchasserName']:
                if datas[d]['PurchasserName'] == '浙江大学' and InvoiceDate < '2015-01-01' and amount <= 1600.00:
                    invoiceInfo.setStatus('通过')
                else:
                    invoiceInfo.setStatus('不通过')
            else:
                invoiceInfo.setStatus('转人工')
        else:
            invoiceInfo.setStatus('转人工')

        insertInvoiceInfo(invoiceInfo)
    print("插入mysql成功")


# 测试函数
def testInsertInvoiceInfo():
    # 测试插入函数
    invoiceInfo = general_invoice_info('2019-01-01', '浙江大学', '浙江大学2', 10000, '通过', 'test')
    insertInvoiceInfo(invoiceInfo)
    # 测试批量插入函数，包括一些异常情况
    datas = []
    data1 = {'InvoiceDate': '2014年01月01日', 'PurchasserName': '浙江大学2', 'SellerName': '浙江大学',
             'AmountInFiguers': '100',
             'invoiceName': 'test1'}
    # 某些字段为空的情况
    data2 = {'InvoiceDate': '2014年01月01日', 'PurchasserName': '浙江大学2', 'SellerName': '浙江大学',
             'AmountInFiguers': '',
             'invoiceName': 'test2'}
    data3 = {'InvoiceDate': '2014年01月01日', 'PurchasserName': '浙江大学2', 'SellerName': '',
             'AmountInFiguers': '100',
             'invoiceName': 'test3'}
    data4 = {'InvoiceDate': '', 'PurchasserName': '浙江大学', 'SellerName': '浙江大学',
             'AmountInFiguers': '100',
             'invoiceName': 'test4'}
    data5 = {'InvoiceDate': '2014年01月01日', 'PurchasserName': '', 'SellerName': '浙江大学2',
             'AmountInFiguers': '100',
             'invoiceName': 'test5'}
    data6 = {'InvoiceDate': '2014年01月01日', 'PurchasserName': '浙江大学2', 'SellerName':'',
             'AmountInFiguers': '100',
             'invoiceName': 'test6'}
    data7 = {'InvoiceDate': '2014年01月01日', 'PurchasserName': '浙江大学2', 'SellerName': '浙江大学',
             'AmountInFiguers': '',
             'invoiceName': 'test7'}
    # 日期不通过的情况
    data8 = {'InvoiceDate': '年月日', 'PurchasserName': '浙江大学2', 'SellerName': '浙江大学',
             'AmountInFiguers': '100',
             'invoiceName': 'test8'}
    data9 = {'InvoiceDate': '2019年01月01日', 'PurchasserName': '浙江大学2', 'SellerName': '浙江大学',
             'AmountInFiguers': '100',
             'invoiceName': 'test9'}

    # 日期通过的情况
    data10 = {'InvoiceDate': '2014年01月01日', 'PurchasserName': '浙江大学', 'SellerName': '浙江大学2',
              'AmountInFiguers': '100',
              'invoiceName': 'test10'}
    # 金额不通过的情况
    data11 = {'InvoiceDate': '2014年01月01日', 'PurchasserName': '浙江大学', 'SellerName': '浙江大学2',
              'AmountInFiguers': '1700',
              'invoiceName': 'test11'}
    # 买方不是浙江大学的情况
    data12 = {'InvoiceDate': '2014年01月01日', 'PurchasserName': '浙江大学1', 'SellerName': '浙江大学2',
              'AmountInFiguers': '100',
              'invoiceName': 'test12'}
    # 把上述数据添加到datas中
    # datas.append(data1)
    # datas.append(data2)
    # datas.append(data3)
    datas.append(data4)
    # datas.append(data5)
    # datas.append(data6)
    # datas.append(data7)
    datas.append(data8)
    datas.append(data9)
    datas.append(data10)
    datas.append(data11)
    datas.append(data12)
    # 批量插入
    insertInvoiceInfoList(datas)


# 查询函数，查询所有PurchaseName不为空，SellerName不为空，status不等于转人工的发票信息
def queryInvoiceInfoOfTransaction():
    sql = "select * from general where PurchasserName != '' and SellerName != '' and status != '转人工'"
    cursor.execute(sql)
    results = cursor.fetchall()
    print("查询成功")
    return results

# 查询函数，查询所有status等于转人工的发票，返回invoiceName
def select_invoice_of_email():
    sql = "SELECT invoiceName FROM general WHERE status = '转人工'"
    cursor.execute(sql)
    results = cursor.fetchall()
    print(list(results))
    print("查询成功")
    return results


# 清空general数据表
def clearTable():
    sql = "truncate table general"
    cursor.execute(sql)
    conn.commit()
    print(" general 表清空成功")


# 统计记录数
def countTable():
    sql = "select count(*) from general"
    cursor.execute(sql)
    results = cursor.fetchall()
    print(" general 表记录数：", results[0][0])
    return results[0][0]


# 定义一个函数做数据统计，统计status是通过的有多少条数据
def count_pass():
    sql = "select count(*) from general where status='通过'"
    cursor.execute(sql)
    result = cursor.fetchall()
    return result[0][0]

# 定义一个函数做数据统计，统计status是不通过的有多少条数据
def count_not_pass():
    # 定义sql语句
    sql = "select count(*) from general where status='不通过'"
    # 执行sql语句
    cursor.execute(sql)
    # 获取查询结果
    result = cursor.fetchall()
    # 打印查询结果
    return result[0][0]

# 定义一个函数做数据统计，统计status是转人工的有多少条数据
def count_to_human():
    # 定义sql语句
    sql = "select count(*) from general where status='转人工'"
    # 执行sql语句
    cursor.execute(sql)
    # 获取查询结果
    result = cursor.fetchall()
    # 打印查询结果
    return result[0][0]

# 导出excel
def export_to_excel():
    # 定义sql语句
    sql = "select * from general order by AmountInFiguers desc"
    # 执行sql语句
    cursor.execute(sql)
    # 获取查询结果
    result = cursor.fetchall()
    # 打印查询结果
    print('result:', result)
    # 提交事务
    conn.commit()
    # 定义一个workbook对象
    wb = openpyxl.Workbook()
    # 定义一个sheet对象
    ws = wb.active
    ws.title = "sheet1"
    # 写入表头
    title = ['发票名称', '开票日期', '购买方名称', '卖方名称', '购买金额', '审批状态', ]
    for i in range(len(title)):
        ws.cell(row=1, column=i + 1, value=title[i])
    # 写入查询结果
    for i in range(len(result)):
        for j in range(len(result[i]) - 2):
            ws.cell(row=i + 2, column=j + 1, value=result[i][j + 1])
    # 写入交易金额和审批状态，交易金额要除以100，并且保留两位小数
    for i in range(len(result)):
        amount = "{:.2f}".format(float(result[i][5]) / 100)
        ws.cell(row=i + 2, column=5, value=amount)
        ws.cell(row=i + 2, column=6, value=result[i][6])

    # add_pie_chart(ws, len(result))
    # 保存数据总条数
    ws.cell(row=1, column=8, value='数据总条数')
    ws.cell(row=2, column=8, value=countTable())
    # 保存excel表格
    wb.save(config.excel_filename_general)


# 删除excel文件
def delete_excel():
    if os.path.exists(config.excel_filename_general):
        os.remove(config.excel_filename_general)
    else:
        print("文件不存在")


#定义一个清空表函数，清空数据库中所有发票信息
def delete_all():
    # 定义sql语句
    sql = "delete from general"
    # 执行sql语句
    cursor.execute(sql)
    # 提交事务
    conn.commit()
    print("general清空成功")


# 测试函数
def testDebug():
    data2 = {'InvoiceDate': '2014年01月01日', 'PurchasserName': '浙江大学2', 'SellerName': '浙江大学',
             'invoiceName': 'test2'}
    data3 = {'InvoiceDate': '2014年01月01日', 'PurchasserName': '浙江大学2', 'AmountInFiguers': '100',
             'invoiceName': 'test3'}
    data = []
    data.append(data2)
    data.append(data3)
    insertInvoiceInfoList(data)


def main():
    # # 清空general数据表
    # delete_excel()
    clearTable()
    # # 测试插入函数
    # invoiceInfo = general_invoice_info('2019-01-01', '浙江大学', '浙江大学2', 10000, '通过', 'test')
    # insertInvoiceInfo(invoiceInfo)
    testInsertInvoiceInfo()
    # testDebug()
    # print(queryInvoiceInfoOfTransaction())
    export_to_excel()


# 定义测试主函数
if __name__ == '__main__':
    main()
